#!/usr/bin/env python3
"""Quick workbook inspection helper for the builtin spreadsheet skill.

Prints workbook structure, sheet visibility, dimensions, and optional row samples.
Requires `openpyxl`.
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path


def sample_csv(path: Path, samples: int) -> int:
    print(f"path: {path}")
    print("format: csv_like")
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        print("sampled_rows:")
        for index, row in enumerate(reader, start=1):
            print(f"  {index}: {row}")
            if index >= samples:
                break
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Inspect a workbook with openpyxl.")
    parser.add_argument("workbook_path", help="Path to the workbook or csv file")
    parser.add_argument("--sheet", help="Optional sheet name to inspect more closely")
    parser.add_argument(
        "--samples",
        type=int,
        default=5,
        help="Number of sampled rows to print from the selected sheet",
    )
    parser.add_argument(
        "--show-formulas",
        action="store_true",
        help="Show formulas instead of values when sampling xlsx rows",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook_path)
    if not workbook_path.is_file():
        print(f"error: file not found: {workbook_path}", file=sys.stderr)
        return 1

    if workbook_path.suffix.lower() in {".csv", ".tsv"}:
        return sample_csv(workbook_path, args.samples)

    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "error: openpyxl is not installed. Install with `uv pip install openpyxl` "
            "or `python3 -m pip install openpyxl`.",
            file=sys.stderr,
        )
        return 2

    try:
        workbook = load_workbook(
            str(workbook_path),
            data_only=not args.show_formulas,
        )
    except Exception as exc:  # pragma: no cover - defensive runtime wrapper
        print(f"error: failed to open workbook: {exc}", file=sys.stderr)
        return 3

    print(f"path: {workbook_path}")
    print(f"format: {workbook_path.suffix.lower().lstrip('.') or 'unknown'}")
    print(f"sheets: {len(workbook.sheetnames)}")
    print("sheet_overview:")
    for name in workbook.sheetnames:
        sheet = workbook[name]
        print(
            f"  - {name}: hidden={sheet.sheet_state != 'visible'} "
            f"rows={sheet.max_row} cols={sheet.max_column}"
        )

    target_name = args.sheet or workbook.sheetnames[0]
    if target_name not in workbook.sheetnames:
        print(f"error: unknown sheet: {target_name}", file=sys.stderr)
        return 4

    sheet = workbook[target_name]
    print(f"selected_sheet: {target_name}")
    print(f"selected_sheet_hidden: {sheet.sheet_state != 'visible'}")
    print("sampled_rows:")
    for index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=args.samples, values_only=True),
        start=1,
    ):
        print(f"  {index}: {row}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
